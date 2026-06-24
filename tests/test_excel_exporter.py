from io import BytesIO

from openpyxl import load_workbook

from excel_exporter import build_excel
from sms_parser import FINAL_COLUMNS


def _row(reason: str) -> dict:
    row = {column: "" for column in FINAL_COLUMNS}
    row["사건번호"] = "2025타경1"
    row["물건번호"] = "1"
    row["확인필요사유"] = reason
    return row


def test_review_sheet_excludes_lookup_not_run_only():
    workbook = load_workbook(BytesIO(build_excel([_row("주소조회 미실행"), _row("전화번호 미확인")], "원문")))
    review_sheet = workbook["확인필요"]

    assert review_sheet.max_row == 2
    assert review_sheet["P2"].value == "전화번호 미확인"


def test_compact_photo_style_sheet_is_included():
    row = _row("")
    row.update({"날짜": "2026-06-17", "법원": "수원", "물건번호": "2", "물건정보": "상가"})
    workbook = load_workbook(BytesIO(build_excel([row], "원문")))
    sheet = workbook["사진형목록"]

    assert sheet["A2"].value == 1
    assert sheet["C2"].value == "6월 17일"
    assert sheet["D2"].value == "25-1 / 2번"
