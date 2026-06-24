from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sms_parser import FINAL_COLUMNS


REVIEW_REASON_KEYWORDS = [
    "전화번호 앞자리 누락 가능성",
    "전화번호 미확인",
    "지번주소 미확인",
    "도로명주소 미확인",
    "법원경매 조회 실패",
    "물건번호 매칭 실패",
    "사건번호 파싱 실패",
    "물건번호 미기재, 1번 추정",
]


def build_excel(rows: list[dict], original_text: str) -> bytes:
    """엑셀 파일 bytes를 반환한다."""
    workbook = Workbook()
    result_sheet = workbook.active
    result_sheet.title = "최종결과"
    _write_table(result_sheet, rows, FINAL_COLUMNS)

    compact_sheet = workbook.create_sheet("사진형목록")
    _write_compact_table(compact_sheet, rows)

    review_rows = [row for row in rows if _needs_review(row)]
    review_sheet = workbook.create_sheet("확인필요")
    _write_table(review_sheet, review_rows, FINAL_COLUMNS)

    original_sheet = workbook.create_sheet("원본문자")
    original_sheet["A1"] = original_text or ""
    original_sheet["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    original_sheet.column_dimensions["A"].width = 100
    original_sheet.row_dimensions[1].height = 300

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _write_compact_table(sheet, rows: list[dict]) -> None:
    columns = ["번호", "법원", "날짜", "사건번호/물건번호", "물건정보", "비고", "전화번호", "전체주소"]
    compact_rows = []
    for index, row in enumerate(rows, start=1):
        case_no = str(row.get("사건번호", ""))
        item_no = str(row.get("물건번호", "")).strip()
        display_case = case_no.replace("타경", "-")
        if len(display_case) >= 4 and display_case[:4].isdigit():
            display_case = display_case[2:]
        if item_no and item_no != "1":
            display_case = f"{display_case} / {item_no}번"
        compact_rows.append(
            {
                "번호": index,
                "법원": row.get("법원", ""),
                "날짜": _display_date(str(row.get("날짜", ""))),
                "사건번호/물건번호": display_case,
                "물건정보": row.get("물건정보", ""),
                "비고": row.get("비고", ""),
                "전화번호": row.get("전화번호", ""),
                "전체주소": row.get("전체주소", "") or row.get("도로명주소", "") or row.get("지번주소", ""),
            }
        )
    _write_table(sheet, compact_rows, columns)


def _display_date(value: str) -> str:
    try:
        parsed = __import__("datetime").date.fromisoformat(value)
        return f"{parsed.month}월 {parsed.day}일"
    except (TypeError, ValueError):
        return value


def _needs_review(row: dict) -> bool:
    reason = str(row.get("확인필요사유", "")).strip()
    if not reason:
        return False
    return any(keyword in reason for keyword in REVIEW_REASON_KEYWORDS)


def _write_table(sheet, rows: list[dict], columns: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    sheet.append(columns)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    sheet.freeze_panes = "A2"
    for column_index, column_name in enumerate(columns, start=1):
        max_length = len(column_name)
        for row_index in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is not None:
                max_length = max(max_length, min(len(str(value)), 60))
        sheet.column_dimensions[get_column_letter(column_index)].width = max(10, max_length + 2)
